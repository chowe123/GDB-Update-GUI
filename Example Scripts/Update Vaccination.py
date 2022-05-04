from __future__ import division
import sys
import os

#Module Folder name
openpyxl_module = "Modules"
module_path = os.path.join(os.path.dirname(__file__), openpyxl_module)
sys.path.append(module_path)

import openpyxl
import arcpy

from datetime import datetime

temp_excel_loc = os.path.join(os.path.dirname(__file__), "Temp", "Temp.xlsx")
excelfile_loc = ""
geodatabase = ""

#Formats the header of a worksheet
def format_header(workbook, sheet_name):
    sheet = workbook[sheet_name]
    merged_cells = map(str, sheet.merged_cells.ranges)
    if "A1:R1" in merged_cells:
        sheet.unmerge_cells('A1:R1')
        sheet.delete_rows(1, 1)
        return 1
    column_names = ['FSA', '00-04', '05-11', '12-17', '18-24', '25-29', '30-34', '35-39', \
                    '40-44', '45-49', '50-54', '55-59', '60-64', '65-69', '70-74', '75-79', '80+', 'Total']
    for c in range(1, 19):
        sheet.cell(row=1, column=c, value=column_names[c-1])
    return 0

#Sums vaccination columns of 5 years old and older in a new column with a header specified by col_name in the 19th column
def sum_5plus_doses(workbook, sheet_name, col_name):
    format_header(workbook, sheet_name)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    sheet.cell(row=1, column=19, value=col_name)
    for r in range(2, max_row):
        total = 0
        for c in range(3, 18):
            total += sheet.cell(row=r, column=c).value
        sheet.cell(row=r, column=19, value=total)

#Formats the header, and creates the 5+ vaccination column and saves it in the specified location
def format_vaccination(file_loc, save_loc):
    workbook = openpyxl.load_workbook(file_loc)
    format_header(workbook, "First Doses")
    format_header(workbook, "Second Doses")
    format_header(workbook, "Third Doses")
    sum_5plus_doses(workbook, "First Doses", "First_f")
    sum_5plus_doses(workbook, "Second Doses", "Second_f")
    sum_5plus_doses(workbook, "Third Doses", "Third_f")
    workbook.save(save_loc)
    return workbook

#Imports the specified excel sheet into the specified file geodatabase
def import_sheet(excel_loc, output_loc, sheet_name):
    arcpy.management.Delete(output_loc)
    arcpy.ExcelToTable_conversion(excel_loc, output_loc, sheet_name)

#Updates the vaccination file geodatabase based on the specified excel file
def update_vaccination():
    #Formats header of exile file, saves it into temporary excel file location
    print("Formating Excel File...")
    workbook = format_vaccination(excelfile_loc, temp_excel_loc)
    #Imports Worksheet as a table in the file geodatabase
    print("Importing Sheets into geodatabase...")
    import_sheet(temp_excel_loc, "First", "First Doses")
    import_sheet(temp_excel_loc, "Second", "Second Doses")
    import_sheet(temp_excel_loc, "Third", "Third Doses")
    print("")
    #Joins the 05_11 vaccination, and five+ vaccination fields
    print("Joining First Doses...")
    arcpy.management.JoinField("COVID_Vaccine_Uptake_FSA", "CFSAUID", "First", "FSA", ["F05_11","First_f"])
    #Updates each field to the newly joined values and calculates Rates
    print("   Updating Total_5plus...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "Total_5plus", "[First_f]", "VB", "")
    print("   Updating Vax5_11...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "Vax5_11", "[F05_11]", "VB", "")
    print("   Updating Dose1_5Plus...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "Dose1_5Plus", "[First_f]", "VB", "")
    print("      Calculating UnVax5_11...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "UnVax5_11", "[POP5_11]-[Vax5_11]", "VB", "")
    print("      Calculating VaxRate...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "VaxRate", "[Total_5plus]/[POP_5Plus]", "VB", "")
    print("      Calculating VaxRate5_11")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "VaxRate5_11", "[Vax5_11]/[POP5_11]", "VB", "")
    #Deletes previously joined values
    print("   Deleting Join Fields of First Doses...")
    arcpy.management.DeleteField("COVID_Vaccine_Uptake_FSA", ["F05_11","First_f"])
    print("")
    print("Joining Second Doses...")
    arcpy.management.JoinField("COVID_Vaccine_Uptake_FSA", "CFSAUID", "Second", "FSA", ["Second_f"])
    print("   Updating Dose2_5Plus...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "Dose2_5Plus", "[Second_f]", "VB", "")
    print("      Calculating VaxRate_D2...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "VaxRate_D2", "[Second_f]/[POP_5Plus]", "VB", "")
    print("   Deleting Join Fields of Second Doses...")
    arcpy.management.DeleteField("COVID_Vaccine_Uptake_FSA", ["Second_f"])
    print("")
    print("Joining Third Doses...")
    arcpy.management.JoinField("COVID_Vaccine_Uptake_FSA", "CFSAUID", "Third", "FSA", ["Third_f"])
    print("   Updating Dose3_5Plus...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "Dose3_5Plus", "[Third_f]", "VB", "")
    print("      Calculating VaxRate_D3...")
    arcpy.CalculateField_management("COVID_Vaccine_Uptake_FSA", "VaxRate_D3", "[Third_f]/[POP_5Plus]", "VB", "")
    print("   Deleting Join Fields of Third Doses...")
    arcpy.management.DeleteField("COVID_Vaccine_Uptake_FSA", ["Third_f"])

    #Deletes imported sheet from the file geodatabase
    print("")
    print("Deleting Imported Sheets from geodatabase...")
    arcpy.management.Delete("First")
    arcpy.management.Delete("Second")
    arcpy.management.Delete("Third")
    print("")
    output_excel = os.path.join(os.path.dirname(__file__), str(datetime.now()).rsplit(".")[0].replace(":", "") + ".xls")
    print("Exporting Updated Vaccination to Excel: {}...".format(output_excel))
    arcpy.conversion.TableToExcel("COVID_Vaccine_Uptake_FSA", output_excel)
    print("")
    print("Completed!")



##################################################################################################
#
#              GUI Code
#
##################################################################################################


from Tkinter import *
import tkFileDialog as filedialog
import tkMessageBox as messagebox

root = Tk()
root.title("Update Vaccination Data")
scrollbar = Scrollbar(orient="horizontal")
fileentry = Entry(root, xscrollcommand=scrollbar.set, width=100)
fileentry.focus()
fileentry.grid(row=0, column=1, sticky="w")
scrollbar.grid(row=1, column=1, ipadx=275, sticky="w")
scrollbar.config(command=fileentry.xview)
fileentry.config()

def selectexcel():
    x = filedialog.askopenfilename(title="Select File")
    datafile = x
    fileentry.delete('0', END)
    fileentry.insert(END, datafile)

dataselect = Button(root, text="Select Excel Sheet", padx=50, command=selectexcel)
dataselect.grid(row=0, column=0)


scrollbar2 = Scrollbar(orient="horizontal")
foldentry = Entry(root, xscrollcommand=scrollbar2.set, width=100)
foldentry.focus()
foldentry.grid(row=2, column=1, sticky="w")
scrollbar2.grid(row=3, column=1, ipadx=275, sticky="w")
scrollbar2.config(command=fileentry.xview)
foldentry.config()

def selectgdb():
    x = filedialog.askdirectory(title="Select GDB")
    datafile = x
    foldentry.delete('0', END)
    foldentry.insert(END, datafile)

dataselect = Button(root, text="Select GDB folder", padx=50, command=selectgdb)
dataselect.grid(row=2, column=0)

def mainrun():
    global excelfile_loc
    global output_loc
    arcpy.env.workspace = foldentry.get()
    excelfile_loc = fileentry.get()
    update_vaccination()
    messagebox.showinfo("Completed", "Vaccine gdb has been updated!")
    

runbutton = Button(root, text="Calculate", command=mainrun)
runbutton.grid(row=4, column=0, ipadx=50, sticky="w")

root.mainloop()
